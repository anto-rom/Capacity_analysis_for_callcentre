
"""
prophet_forecast_v6.py
Version 6 – Enhancements for Language Staffing + Full pipeline
- Prophet on Net_Emails + weekend correction + dynamic RR by weekday (trend-aware)
- Shrinkage-aware capacity scenarios (WD 18/20/24; weekend 7) & charts
- Language staffing:
  * Default base: Total_Workload_forecast (configurable to Net if desired)
  * Agents_lang = ceil(Workload_lang / CASES_PER_AGENT)
  * Scenario Picker (18/20/24 Mon–Fri) with language apportionment via “Largest Remainders”
  * Language charts (next 30 days and weekly Thu–Wed)
  * “Sanity” sheet with an explicit check for the next date (e.g., English)
Inputs
- Items: C:/Users/pt3canro/Desktop/CAPACITY/items_per_day.xlsx
- Shrinkage: C:/Users/pt3canro/Desktop/CAPACITY/Historical Shrinkage.xlsx
Outputs
- Main Excel: C:/Users/pt3canro/Desktop/CAPACITY/OUTPUTS/future_preditions.xlsx
- Language staffing: C:/Users/pt3canro/Desktop/CAPACITY/OUTPUTS/language_staffing_requirements.xlsx
- Charts PNG saved in the same OUTPUTS folder and inserted into the Excels
Dependencies: prophet, pandas, numpy, openpyxl, matplotlib, seaborn
"""
import os
import math
import warnings
from dataclasses import dataclass
from typing import Dict, Tuple, List, Optional
import numpy as np
import pandas as pd
from prophet import Prophet
# Plotting and embedding charts into Excel
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.dates import AutoDateLocator, ConciseDateFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
warnings.filterwarnings("ignore", category=FutureWarning)

# =========================== CONFIG ===========================
@dataclass
class Config:
    # Paths
    PATH_ITEMS: str = r"C:/Users/pt3canro/Desktop/CAPACITY/items_per_day.xlsx"
    PATH_SHRINK: str = r"C:/Users/pt3canro/Desktop/CAPACITY/Historical Shrinkage.xlsx"
    OUTPUT_XLSX: str = r"C:/Users/pt3canro/Desktop/CAPACITY/OUTPUTS/future_preditions.xlsx"
    OUTPUT_LANG_XLSX: str = r"C:/Users/pt3canro/Desktop/CAPACITY/OUTPUTS/language_staffing_requirements.xlsx"
    # Forecast horizon
    HORIZON_DAYS: int = 365
    # Productivity (cases per agent)
    CASES_PER_AGENT: int = 18
    # Staffing
    WEEKEND_AGENTS: int = 7
    WEEKDAY_AGENT_OPTIONS: Tuple[int, int, int] = (18, 20, 24)
    # Minimum lookback (>= 16 weeks)
    LOOKBACK_MIN_DAYS: int = 112
    # Weekend factors bounds and fallbacks
    SAT_FACTOR_BOUNDS: Tuple[float, float] = (0.40, 0.90)
    SUN_FACTOR_BOUNDS: Tuple[float, float] = (0.40, 0.90)
    SAT_FACTOR_FALLBACK: float = 0.53
    SUN_FACTOR_FALLBACK: float = 0.60
    # Prophet
    YEARLY_SEASONALITY: bool = True
    WEEKLY_SEASONALITY: int = 10
    SEASONALITY_MODE: str = "multiplicative"
    CHANGEPOINT_PRIOR: float = 0.05
    INTERVAL_WIDTH: float = 0.80
    # Language mix (percentage)
    LANG_SPLIT_PCT: Dict[str, float] = None
    # Base for language staffing: 'total' (Total_Workload_forecast) or 'net' (yhat_net_adj)
    LANG_BASE: str = 'total'
    # Business rules: date-based multipliers applied ONLY to future forecasts
    # Format: [(start_date, end_date, multiplier), ...]
    DATE_MULTIPLIERS: List[Tuple[str, str, float]] = None

CFG = Config()
CFG.LANG_SPLIT_PCT = {
    'English': 64.35,
    'French': 7.41,
    'German': 8.60,
    'Italian': 6.67,
    'Portuguese': 1.62,
    'Spanish': 11.35,
}
# >>> Business rules you requested (applied ONLY to 'future'):
#     - +7% from 2025-09-01 to 2025-09-30
#     - -3% from 2025-10-01 to 2025-12-31 (Q4 reduction)
CFG.DATE_MULTIPLIERS = [
    ("2025-09-01", "2025-09-30", 1.07),  # +7% Sep 2025 only
    ("2025-10-01", "2026-04-30", 0.95),  # -5% Oct-Dec 2025 only
]

# =========================== UTILITIES ===========================
def _assert_columns(df: pd.DataFrame, must_have: List[str], fname: str):
    missing = [c for c in must_have if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in {fname}: {missing}")

def _read_items(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Not found: {path}")
    df = pd.read_excel(path)
    _assert_columns(df, ['Date', 'Emails', 'Spam', 'Outage', 'RR'], 'items_per_day.xlsx')
    df['Date'] = pd.to_datetime(df['Date'])
    for c in ['Emails', 'Spam', 'Outage', 'RR']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df['Net_Emails'] = (df['Emails'] - df['Spam'] - df['Outage']).clip(lower=0)
    df['Total_Workload_real'] = df['Net_Emails'] * (1.0 + (df['RR'].fillna(0) / 100.0))
    df['Weekday'] = df['Date'].dt.day_name()
    return df.sort_values('Date').reset_index(drop=True)

def _read_shrinkage(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame(columns=['Date', 'weekday', 'shrinkage_hours'])
    df = pd.read_excel(path)
    date_col = None
    for cand in ['Date', 'date', 'FECHA']:
        if cand in df.columns:
            date_col = cand
            break
    if date_col is None:
        raise ValueError("Historical Shrinkage.xlsx needs a Date column")
    if 'shrinkage hours' in df.columns:
        df['shrinkage_hours'] = pd.to_numeric(df['shrinkage hours'], errors='coerce')
    elif 'shrinkage_hours' in df.columns:
        df['shrinkage_hours'] = pd.to_numeric(df['shrinkage_hours'], errors='coerce')
    elif 'shrinkage seconds' in df.columns:
        df['shrinkage_hours'] = pd.to_numeric(df['shrinkage seconds'], errors='coerce') / 3600.0
    elif 'shrinkage_seconds' in df.columns:
        df['shrinkage_hours'] = pd.to_numeric(df['shrinkage_seconds'], errors='coerce') / 3600.0
    else:
        return pd.DataFrame(columns=['Date', 'weekday', 'shrinkage_hours'])
    df['Date'] = pd.to_datetime(df[date_col])
    df['weekday'] = df['Date'].dt.day_name()
    return df[['Date', 'weekday', 'shrinkage_hours']].dropna().sort_values('Date').reset_index(drop=True)

def _fit_prophet_on_net(net_df: pd.DataFrame):
    prophet_df = net_df[['Date', 'Net_Emails']].rename(columns={'Date': 'ds', 'Net_Emails': 'y'})
    m = Prophet(
        yearly_seasonality=CFG.YEARLY_SEASONALITY,
        weekly_seasonality=CFG.WEEKLY_SEASONALITY,
        seasonality_mode=CFG.SEASONALITY_MODE,
        changepoint_prior_scale=CFG.CHANGEPOINT_PRIOR,
        interval_width=CFG.INTERVAL_WIDTH,
    )
    m.fit(prophet_df)
    future = m.make_future_dataframe(periods=CFG.HORIZON_DAYS, freq='D')
    forecast = m.predict(future)
    return m, forecast, prophet_df

# --------------- RR: weekday-specific, trend-aware ----------------
def _rr_dynamic_by_dow(items: pd.DataFrame) -> Dict[str, float]:
    """
    Estimate RR% per weekday using a trend-aware (EWM) level per weekday.
    Logic:
    - For each weekday, compute an exponentially weighted mean (EWM) over the
      sub-series of that weekday only (i.e., Mondays use only Mondays), with a
      ~4-week halflife to capture recent trend while being robust to noise.
    - If a weekday has insufficient history, fall back to:
      RR_hat(dow) = RR_level_global_EWM * median(dow) / median(global)
      which retains weekday seasonality relative to the global level.
    - Always return a dict with all 7 weekdays.
    Returns
    ----
    Dict[str, float] : {'Monday': rr_hat, ..., 'Sunday': rr_hat}
    """
    dows = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    df = items.dropna(subset=['RR']).copy()
    if df.empty:
        return {dow: 0.0 for dow in dows}
    df['weekday'] = df['Date'].dt.day_name()
    rr_series = df.set_index('Date')['RR'].astype(float).sort_index()
    # Global EWM level for fallback (28-day halflife similar to previous version)
    rr_level_global = float(
        rr_series.ewm(
            halflife=28,
            min_periods=min(CFG.LOOKBACK_MIN_DAYS, max(1, len(rr_series)//4))
        ).mean().iloc[-1]
    )
    med_global = float(df['RR'].median()) if df['RR'].notna().any() else rr_level_global
    med_by_dow = df.groupby('weekday')['RR'].median().to_dict()
    rr_hat_by_dow: Dict[str, float] = {}
    # Weekly cadence for each weekday sub-series → ~4-week halflife
    weekly_halflife = 4.0
    for dow in dows:
        sub = df[df['weekday'] == dow].sort_values('Date')
        if len(sub) >= 4:  # enough weekly points to compute a stable EWM
            rr_level_dow = float(
                sub.set_index('Date')['RR']
                .ewm(halflife=weekly_halflife, min_periods=max(2, len(sub)//3))
                .mean().iloc[-1]
            )
            rr_hat = max(0.0, rr_level_dow)
        else:
            # Fallback keeps weekday-to-global ratio
            ratio = (med_by_dow.get(dow, med_global)) / (med_global if med_global != 0 else 1.0)
            rr_hat = max(0.0, rr_level_global * ratio)
        rr_hat_by_dow[dow] = rr_hat
    return rr_hat_by_dow

def _weekend_factors(hist_join: pd.DataFrame) -> Dict[str, float]:
    df = hist_join.copy()
    df.loc[:, 'weekday'] = df['ds'].dt.day_name()
    def _calc(dow: str, bounds: Tuple[float, float], fb: float) -> float:
        sub = df[(df['weekday'] == dow) & df['Net_Emails'].notna() & df['yhat_net'].notna()].copy()
        if len(sub) >= 4:
            ratios = (sub['Net_Emails'] / sub['yhat_net']).replace([np.inf, -np.inf], np.nan).dropna()
            f = float(ratios.median()) if not ratios.empty else fb
        else:
            f = fb
        return max(bounds[0], min(bounds[1], f))
    sat = _calc('Saturday', CFG.SAT_FACTOR_BOUNDS, CFG.SAT_FACTOR_FALLBACK)
    sun = _calc('Sunday', CFG.SUN_FACTOR_BOUNDS, CFG.SUN_FACTOR_FALLBACK)
    factors = {dow: 1.0 for dow in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']}
    factors['Saturday'] = sat
    factors['Sunday'] = sun
    return factors

def _apply_weekend_factor(df: pd.DataFrame, factors: Dict[str, float]):
    return df['Weekday'].map(factors).fillna(1.0)

def _build_capacity_backlog_with_shrink(future_df: pd.DataFrame, shrink_med_by_dow: Dict[str, float]) -> pd.DataFrame:
    out = future_df.copy()
    out.loc[:, 'Shrinkage_hours_est'] = out['Weekday'].map(shrink_med_by_dow).fillna(0.0)
    out.loc[:, 'Lost_Agents_Equiv'] = (out['Shrinkage_hours_est'] / 8.0).clip(lower=0.0)
    def _scenario_cols(weekday_agents: int) -> Tuple[str, str, str, str]:
        return (
            f"Agents_WD{weekday_agents}",
            f"AgentsEff_WD{weekday_agents}",
            f"Capacity_WD{weekday_agents}",
            f"CumBacklog_WD{weekday_agents}",
        )
    for opt in CFG.WEEKDAY_AGENT_OPTIONS:
        col_a, col_ae, col_c, col_b = _scenario_cols(opt)
        sched = np.where(out['Weekday'].isin(['Saturday', 'Sunday']), CFG.WEEKEND_AGENTS, opt)
        out.loc[:, col_a] = sched.astype(float)
        out.loc[:, col_ae] = (out[col_a] - out['Lost_Agents_Equiv']).clip(lower=0.0)
        out.loc[:, col_c] = (out[col_ae] * CFG.CASES_PER_AGENT).round(0)
        # Deterministic backlog (carry positive remainder forward)
        backlog = 0.0
        backs: List[float] = []
        for tw, cap in zip(out['Total_Workload_forecast'].astype(float), out[col_c].astype(float)):
            daily_backlog = tw - cap
            backlog = max(0.0, backlog + daily_backlog)
            backs.append(backlog)
        out.loc[:, col_b] = np.array(backs, dtype=float)
    return out

# =========================== PLOTS ===========================
def _plot_avg_real_by_weekday(hist: pd.DataFrame, out_path: str):
    weekday_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    df = (
        hist.dropna(subset=['Total_Workload_real'])
        .groupby('Weekday', as_index=False)['Total_Workload_real']
        .mean()
        .set_index('Weekday')
        .reindex(weekday_order)
    )
    plt.figure(figsize=(10, 5))
    ax = sns.barplot(x=df.index, y=df['Total_Workload_real'].values, color="#4C78A8")
    ax.set_title('Average Total Workload (Real) by Weekday', fontsize=12)
    ax.set_xlabel('Weekday')
    ax.set_ylabel('Avg Total Workload')
    plt.xticks(rotation=45)
    ymax = float(np.nanmax(df['Total_Workload_real'].values)) if len(df) > 0 else 0
    for p in ax.patches:
        h = p.get_height()
        ax.text(p.get_x() + p.get_width() / 2.0, h + (0.01 * max(1, ymax)), f'{h:.1f}',
                ha='center', va='bottom', fontsize=9)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()

def _plot_avg_real_vs_pred_by_weekday(hist: pd.DataFrame, out_path: str):
    weekday_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    agg = (
        hist.groupby('Weekday', as_index=False)
        .agg(Avg_Real=('Total_Workload_real', 'mean'), Avg_Pred=('Total_Workload_pred', 'mean'))
        .set_index('Weekday')
        .reindex(weekday_order)
        .reset_index()
    )
    plot_df = agg.melt(id_vars='Weekday', value_vars=['Avg_Real', 'Avg_Pred'],
                       var_name='Series', value_name='Avg_Total_Workload')
    plt.figure(figsize=(11, 5))
    ax = sns.barplot(
        data=plot_df,
        x='Weekday', y='Avg_Total_Workload', hue='Series',
        palette={'Avg_Real': '#4C78A8', 'Avg_Pred': '#F58518'}
    )
    ax.set_title('Average Total Workload: Real vs Pred by Weekday', fontsize=12)
    ax.set_xlabel('Weekday')
    ax.set_ylabel('Avg Total Workload')
    plt.xticks(rotation=45)
    if hasattr(ax, 'bar_label'):
        for container in ax.containers:
            ax.bar_label(container, fmt='%.1f', fontsize=8, padding=2)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()

def _plot_daily_trend_real_vs_pred(hist: pd.DataFrame, future: pd.DataFrame, out_path: str):
    plt.figure(figsize=(14, 5))
    ax = plt.gca()
    locator = AutoDateLocator()
    formatter = ConciseDateFormatter(locator)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(formatter)
    mask_real = hist['Total_Workload_real'].notna()
    ax.plot(hist.loc[mask_real, 'ds'], hist.loc[mask_real, 'Total_Workload_real'],
            label='Real', color='#4C78A8', linewidth=1.2)
    mask_pred = hist['Total_Workload_pred'].notna()
    ax.plot(hist.loc[mask_pred, 'ds'], hist.loc[mask_pred, 'Total_Workload_pred'],
            label='Pred (hist)', color='#F58518', linewidth=1.2, alpha=0.85)
    ax.plot(future['ds'], future['Total_Workload_forecast'],
            label='Forecast', color='#F58518', linewidth=1.2, linestyle='--')
    ax.set_title('Daily Trend: Real vs Pred (+Forecast)', fontsize=12)
    ax.set_xlabel('Date')
    ax.set_ylabel('Total Workload')
    ax.grid(True, linestyle='--', alpha=0.3)
    ax.legend(loc='upper left', ncol=3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()

def _plot_lang_agents_next30(daily_lang: pd.DataFrame, langs: List[str], out_path: str):
    # (Function body presumably exists in your original; kept as-is in your file)
    df = daily_lang.copy()
    df = df.sort_values('ds').head(30)
    df = df[['ds'] + [f'Agents_{l}' for l in langs]]
    df = df.set_index('ds')
    plt.figure(figsize=(14, 6))
    ax = plt.gca()
    df.plot(kind='bar', stacked=True, ax=ax, colormap='tab20')
    # Add totals above each bar
    for idx, row in enumerate(df.values):
        total = row.sum()
        ax.text(idx, total + 0.5, str(int(total)), ha='center', va='bottom', fontsize=8)
    plt.title('Agents Required by Language – Next 30 Days (stacked)')
    plt.xlabel('Date')
    plt.ylabel('Agents')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()

def _plot_lang_agents_weekly(weekly_lang: pd.DataFrame, langs: List[str], out_path: str):
    df = weekly_lang[['AWeek_Start'] + [f'Agents_{l}' for l in langs]].copy().sort_values('AWeek_Start')
    df = df.set_index('AWeek_Start')
    plt.figure(figsize=(14, 6))
    df.plot(kind='bar', stacked=True, ax=plt.gca(), colormap='tab20')
    plt.title('Agents Required by Language – Weekly (Thu–Wed)')
    plt.xlabel('Week Start (Thu)')
    plt.ylabel('Agents (sum over week)')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close()

# =========================== LANGUAGE LOGIC / SCENARIO ===========================
def _normalize_lang_split(lang_split_pct: Dict[str, float]) -> Dict[str, float]:
    dec = {k: float(v) / 100.0 for k, v in lang_split_pct.items()}
    s = sum(dec.values())
    if s <= 0:
        raise ValueError('Language split percentages must sum > 0')
    if not (0.999 <= s <= 1.001):
        dec = {k: v / s for k, v in dec.items()}
    return dec

def _antonio_week_start(d: pd.Series) -> pd.Series:
    # Week is Thu .. Wed → return the THURSDAY of each week
    return d - pd.to_timedelta((d.dt.dayofweek - 3) % 7, unit='D')

def _largest_remainders_apportion(total_agents: int, weights: Dict[str, float]) -> Dict[str, int]:
    if total_agents <= 0:
        return {k: 0 for k in weights}
    s = sum(weights.values()) or 1.0
    w = {k: v / s for k, v in weights.items()}
    quotas = {k: total_agents * w[k] for k in w}
    floors = {k: int(np.floor(quotas[k])) for k in w}
    assigned = sum(floors.values())
    remaining = total_agents - assigned
    if remaining > 0:
        frac = sorted([(k, quotas[k] - floors[k]) for k in w], key=lambda x: x[1], reverse=True)
        for i in range(remaining):
            floors[frac[i % len(frac)][0]] += 1
    return floors

def build_language_staffing(base_df: pd.DataFrame, cases_per_agent: int, lang_split_pct: Dict[str, float], base_col: str):
    """
    base_col: 'Total_Workload_forecast' or 'yhat_net_adj'
    """
    lang_dec = _normalize_lang_split(lang_split_pct)
    langs = list(lang_dec.keys())
    daily = base_df[['ds', 'Weekday', base_col]].copy()
    daily = daily.rename(columns={base_col: 'Workload_Base'})
    # Split + agents per language/day (ceil)
    for lang in langs:
        daily[f'Workload_{lang}'] = daily['Workload_Base'] * lang_dec[lang]
        daily[f'Agents_{lang}'] = np.ceil(daily[f'Workload_{lang}'] / cases_per_agent).astype(int)
    # Global controls
    daily['Agents_Global_Ceil'] = np.ceil(daily['Workload_Base'] / cases_per_agent).astype(int)
    daily['Agents_ByLang_Sum'] = daily[[f'Agents_{l}' for l in langs]].sum(axis=1)
    daily['Rounding_Overhead'] = daily['Agents_ByLang_Sum'] - daily['Agents_Global_Ceil']
    # Weekly aggregation (Thu–Wed)
    daily['AWeek_Start'] = _antonio_week_start(daily['ds'])
    agg_work = {f'Workload_{l}': 'sum' for l in langs}
    weekly = daily.groupby('AWeek_Start', as_index=False).agg({**{'Workload_Base': 'sum'}, **agg_work})
    for lang in langs:
        weekly[f'Agents_{lang}'] = np.ceil(weekly[f'Workload_{lang}'] / cases_per_agent).astype(int)
    weekly['Agents_Global_Ceil'] = np.ceil(weekly['Workload_Base'] / cases_per_agent).astype(int)
    weekly['Agents_ByLang_Sum'] = weekly[[f'Agents_{l}' for l in langs]].sum(axis=1)
    weekly['Rounding_Overhead'] = weekly['Agents_ByLang_Sum'] - weekly['Agents_Global_Ceil']
    # Meta
    meta_rows = [(k, f"{lang_dec[k] * 100:.2f}%") for k in langs]
    meta_rows += [("LANG_BASE", base_col)]
    meta = pd.DataFrame(meta_rows, columns=['Key', 'Value'])
    return daily, weekly, meta, lang_dec, langs

def _load_existing_scenario(path: str) -> Optional[pd.DataFrame]:
    if not os.path.exists(path):
        return None
    try:
        df = pd.read_excel(path, sheet_name='Scenario_Picker')
        if {'ds', 'Scenario_WD'}.issubset(df.columns):
            df['ds'] = pd.to_datetime(df['ds'])
            return df[['ds', 'Scenario_WD']]
    except Exception:
        return None
    return None

def build_scenario_picker(future_df: pd.DataFrame, existing: Optional[pd.DataFrame]) -> pd.DataFrame:
    sp = future_df[['ds', 'Weekday']].copy()
    sp['Scenario_WD'] = np.where(sp['Weekday'].isin(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']), 18, np.nan)
    if existing is not None and not existing.empty:
        sp = sp.merge(existing, on='ds', how='left', suffixes=('', '_old'))
        sp['Scenario_WD'] = np.where(sp['Scenario_WD_old'].notna(), sp['Scenario_WD_old'], sp['Scenario_WD'])
        sp = sp.drop(columns=['Scenario_WD_old'])
    return sp

def build_daily_comparison(daily_need: pd.DataFrame, scenario_picker: pd.DataFrame, lang_weights: Dict[str, float], langs: List[str], weekend_agents: int) -> pd.DataFrame:
    df = daily_need[['ds', 'Weekday'] + [f'Agents_{l}' for l in langs]].copy()
    df = df.merge(scenario_picker[['ds', 'Scenario_WD']], on='ds', how='left')
    is_weekend = df['Weekday'].isin(['Saturday', 'Sunday'])
    df['Planned_Total'] = np.where(is_weekend, weekend_agents, df['Scenario_WD'].fillna(18)).astype(float)
    # Language apportionment per day using Largest Remainders
    for idx, row in df.iterrows():
        total = int(row['Planned_Total']) if not np.isnan(row['Planned_Total']) else 0
        alloc = _largest_remainders_apportion(total, lang_weights)
        for lang in langs:
            df.loc[idx, f'AgentsPlanned_{lang}'] = alloc[lang]
            df.loc[idx, f'Delta_{lang}'] = alloc[lang] - int(row[f'Agents_{lang}'])
    df['Agents_Required_Total'] = df[[f'Agents_{l}' for l in langs]].sum(axis=1)
    df['Agents_Planned_Total'] = df[[f'AgentsPlanned_{l}' for l in langs]].sum(axis=1)
    df['Delta_Total'] = df['Agents_Planned_Total'] - df['Agents_Required_Total']
    return df

def build_weekly_comparison(daily_comp: pd.DataFrame, langs: List[str]) -> pd.DataFrame:
    tmp = daily_comp.copy()
    tmp['AWeek_Start'] = _antonio_week_start(tmp['ds'])
    agg = {f'Agents_{l}': 'sum' for l in langs}
    agg.update({f'AgentsPlanned_{l}': 'sum' for l in langs})
    agg.update({f'Delta_{l}': 'sum' for l in langs})
    agg.update({'Agents_Required_Total': 'sum', 'Agents_Planned_Total': 'sum', 'Delta_Total': 'sum'})
    weekly = tmp.groupby('AWeek_Start', as_index=False).agg(agg)
    return weekly

# =========================== MAIN ===========================
def main():
    # 1) Load inputs
    items = _read_items(CFG.PATH_ITEMS)
    shrink = _read_shrinkage(CFG.PATH_SHRINK)

    # 2) Prophet on Net_Emails
    m, forecast, prophet_df = _fit_prophet_on_net(items)

    # 3) Historical + predictions on Net
    hist = forecast[['ds', 'yhat']].rename(columns={'yhat': 'yhat_net'})
    hist = hist.merge(items.rename(columns={'Date': 'ds'}), on='ds', how='left').sort_values('ds')

    # 4) Weekend factors
    weekend_factors = _weekend_factors(hist)

    # 5) Dynamic RR by weekday (trend-aware)
    rr_hat_by_dow = _rr_dynamic_by_dow(items)

    # 6) Historical audit (does NOT apply date multipliers; keep history unchanged)
    hist = hist.copy()
    hist.loc[:, 'Weekday'] = hist['ds'].dt.day_name()
    hist.loc[:, 'Weekend_Factor'] = _apply_weekend_factor(hist, weekend_factors)
    hist.loc[:, 'yhat_net_adj'] = hist['yhat_net'] * hist['Weekend_Factor']
    # Map weekday-specific RR_hat (no global fillna)
    hist.loc[:, 'RR_hat'] = hist['Weekday'].map(rr_hat_by_dow)
    hist.loc[:, 'Total_Workload_pred'] = hist['yhat_net_adj'] * (1.0 + hist['RR_hat'] / 100.0)
    hist.loc[:, 'error_TW'] = hist['Total_Workload_real'] - hist['Total_Workload_pred']
    hist.loc[:, 'abs_error_TW'] = hist['error_TW'].abs()
    hist.loc[:, 'mape_TW'] = np.where(
        hist['Total_Workload_real'] > 0,
        (hist['abs_error_TW'] / hist['Total_Workload_real']) * 100.0,
        np.nan
    )

    # 7) Future
    last_hist_date = items['Date'].max()
    future = forecast[forecast['ds'] > last_hist_date][['ds', 'yhat']].rename(columns={'yhat': 'yhat_net'}).copy()
    future.loc[:, 'Weekday'] = future['ds'].dt.day_name()
    future.loc[:, 'Weekend_Factor'] = _apply_weekend_factor(future, weekend_factors)
    future.loc[:, 'yhat_net_adj'] = future['yhat_net'] * future['Weekend_Factor']
    # Map weekday-specific RR_hat (no global fillna)
    future.loc[:, 'RR_hat'] = future['Weekday'].map(rr_hat_by_dow)
    future.loc[:, 'Total_Workload_forecast'] = future['yhat_net_adj'] * (1.0 + future['RR_hat'] / 100.0)

    # 7.3) Apply date-based multipliers to FUTURE ONLY (business rules; non-annual)
    #      This preserves history and limits rules to the specified 2025 windows.
    if CFG.DATE_MULTIPLIERS:
        for start_str, end_str, mult in CFG.DATE_MULTIPLIERS:
            d0 = pd.Timestamp(start_str)
            d1 = pd.Timestamp(end_str)
            mask = (future['ds'] >= d0) & (future['ds'] <= d1)
            future.loc[mask, 'Total_Workload_forecast'] *= float(mult)

    # 8) Shrinkage (median by weekday if provided)
    if not shrink.empty:
        shrink_med_by_dow = shrink.groupby('weekday')['shrinkage_hours'].median().to_dict()
    else:
        shrink_med_by_dow = {}

    # 9) Capacity + backlog scenarios with shrinkage
    future_cap = _build_capacity_backlog_with_shrink(future, shrink_med_by_dow)

    # 10) Main export + charts
    audit_cols = [
        'ds', 'Weekday', 'Net_Emails', 'RR', 'Total_Workload_real',
        'yhat_net', 'Weekend_Factor', 'yhat_net_adj', 'RR_hat', 'Total_Workload_pred',
        'error_TW', 'abs_error_TW', 'mape_TW',
    ]
    hist_out = hist[audit_cols].copy()

    base_cols = [
        'ds', 'Weekday', 'yhat_net', 'Weekend_Factor', 'yhat_net_adj', 'RR_hat', 'Total_Workload_forecast',
        'Shrinkage_hours_est', 'Lost_Agents_Equiv',
    ]
    scenario_cols: List[str] = []
    for opt in CFG.WEEKDAY_AGENT_OPTIONS:
        scenario_cols += [f"Agents_WD{opt}", f"AgentsEff_WD{opt}", f"Capacity_WD{opt}", f"CumBacklog_WD{opt}"]
    future_out = future_cap[base_cols + scenario_cols].copy()

    os.makedirs(os.path.dirname(CFG.OUTPUT_XLSX), exist_ok=True)
    with pd.ExcelWriter(CFG.OUTPUT_XLSX, engine='openpyxl') as writer:
        future_out.to_excel(writer, sheet_name='Forecast', index=False)
        hist_out.to_excel(writer, sheet_name='Audit', index=False)

        # Meta sheet
        meta_rows = [
            ["CASES_PER_AGENT", CFG.CASES_PER_AGENT],
            ["WEEKEND_AGENTS", CFG.WEEKEND_AGENTS],
            ["WEEKDAY_AGENT_OPTIONS", ",".join(map(str, CFG.WEEKDAY_AGENT_OPTIONS))],
            ["LOOKBACK_POLICY", f">= {CFG.LOOKBACK_MIN_DAYS} days; use all history if available"],
            ["SAT_FACTOR", weekend_factors.get('Saturday', 1.0)],
            ["SUN_FACTOR", weekend_factors.get('Sunday', 1.0)],
        ]
        for dow, v in _rr_dynamic_by_dow(items).items():
            meta_rows.append([f"RR_hat_{dow}", v])
        for dow in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
            meta_rows.append([f"Shrinkage_median_hours_{dow}", shrink_med_by_dow.get(dow, 0.0)])
        # (Optional) You can append DATE_MULTIPLIERS here for traceability if desired.
        meta = pd.DataFrame(meta_rows, columns=['Key', 'Value'])
        meta.to_excel(writer, sheet_name='Meta', index=False)

    # Charts (and embed into Excel)
    out_dir = os.path.dirname(CFG.OUTPUT_XLSX)
    fig1_path = os.path.join(out_dir, 'avg_total_workload_by_weekday.png')
    fig2_path = os.path.join(out_dir, 'avg_real_vs_pred_by_weekday.png')
    fig3_path = os.path.join(out_dir, 'daily_trend_real_vs_pred.png')
    try:
        _plot_avg_real_by_weekday(hist, fig1_path)
        _plot_avg_real_vs_pred_by_weekday(hist, fig2_path)
        _plot_daily_trend_real_vs_pred(hist, future, fig3_path)
        wb = load_workbook(CFG.OUTPUT_XLSX)
        ws_meta = wb['Meta'] if 'Meta' in wb.sheetnames else wb.active
        ws_fore = wb['Forecast'] if 'Forecast' in wb.sheetnames else wb.active
        img1 = XLImage(fig1_path); img1.anchor = 'E2'; ws_meta.add_image(img1)
        img2 = XLImage(fig2_path); img2.anchor = 'E22'; ws_meta.add_image(img2)
        img3 = XLImage(fig3_path); img3.anchor = 'R2'; ws_fore.add_image(img3)
        wb.save(CFG.OUTPUT_XLSX)
    except Exception as e:
        print(f"Chart embedding warning (main): {e}")

    # 11) Language staffing
    lang_base_col = 'Total_Workload_forecast' if CFG.LANG_BASE.lower() == 'total' else 'yhat_net_adj'
    base_df = future_out[['ds', 'Weekday', 'Total_Workload_forecast', 'yhat_net_adj']].copy()
    daily_lang, weekly_lang, meta_lang, lang_weights, langs = build_language_staffing(
        base_df, CFG.CASES_PER_AGENT, CFG.LANG_SPLIT_PCT, base_col=lang_base_col
    )

    # Scenario Picker
    existing_sp = _load_existing_scenario(CFG.OUTPUT_LANG_XLSX)
    scenario_picker = build_scenario_picker(future_out[['ds', 'Weekday']], existing_sp)

    # Comparisons
    daily_comp = build_daily_comparison(daily_lang, scenario_picker, lang_weights, langs, CFG.WEEKEND_AGENTS)
    weekly_comp = build_weekly_comparison(daily_comp, langs)

    # 11.1 SANITY CHECK (next date) → "Sanity" sheet and add notes to Meta
    next_date = daily_lang['ds'].min()
    sanity_rows = []
    if pd.notna(next_date):
        # Use English as example
        share_eng = lang_weights.get('English', 0.0)
        row_next = daily_lang.loc[daily_lang['ds'] == next_date].iloc[0]
        total_next = float(row_next['Workload_Base'])
        eng_cases = float(row_next['Workload_English']) if 'Workload_English' in row_next else total_next * share_eng
        eng_agents = int(np.ceil(eng_cases / CFG.CASES_PER_AGENT))
        global_agents = int(np.ceil(total_next / CFG.CASES_PER_AGENT))
        # Sanity rows
        sanity_rows = [
            ['NextDate', str(next_date.date())],
            ['LANG_BASE', lang_base_col],
            ['English_Share_pct', f"{share_eng*100:.2f}%"],
            ['Total_Workload_nextDate', round(total_next, 2)],
            ['Workload_English_nextDate', round(eng_cases, 2)],
            ['Agents_English_nextDate', eng_agents],
            ['Agents_Global_Ceil_nextDate', global_agents]
        ]
        # Also add to meta_lang
        for k, v in sanity_rows:
            meta_lang.loc[len(meta_lang)] = [k, v]
    sanity_df = pd.DataFrame(sanity_rows, columns=['Key', 'Value']) if sanity_rows else pd.DataFrame(columns=['Key','Value'])

    # Language charts
    fig4_path = os.path.join(out_dir, 'lang_agents_next30.png')
    fig5_path = os.path.join(out_dir, 'lang_agents_weekly.png')
    try:
        _plot_lang_agents_next30(daily_lang, langs, fig4_path)
        _plot_lang_agents_weekly(weekly_lang, langs, fig5_path)
    except Exception as e:
        print(f"Chart build warning (language): {e}")

    # Save language staffing Excel
    # Output structure:
    # - Daily_Language_Staffing: Daily agent requirements per language.
    # - Weekly_ThuWed_Summary: Weekly (Thu–Wed) agent requirements per language.
    # - Scenario_Picker: Manual selection of weekday agent scenarios (18/20/24).
    # - Daily_Comparison: Required vs planned agents per language, per day.
    # - Weekly_Comparison: Weekly summary of required vs planned agents per language.
    # - Meta: Metadata and configuration used for calculations.
    # - Sanity: Sanity check for the next forecasted date (if available).
    os.makedirs(os.path.dirname(CFG.OUTPUT_LANG_XLSX), exist_ok=True)
    with pd.ExcelWriter(CFG.OUTPUT_LANG_XLSX, engine='openpyxl') as writer:
        daily_lang.to_excel(writer, sheet_name='Daily_Language_Staffing', index=False)
        weekly_lang.to_excel(writer, sheet_name='Weekly_ThuWed_Summary', index=False)
        scenario_picker.to_excel(writer, sheet_name='Scenario_Picker', index=False)
        daily_comp.to_excel(writer, sheet_name='Daily_Comparison', index=False)
        weekly_comp.to_excel(writer, sheet_name='Weekly_Comparison', index=False)
        meta_lang.to_excel(writer, sheet_name='Meta', index=False)
        if not sanity_df.empty:
            sanity_df.to_excel(writer, sheet_name='Sanity', index=False)

    # Embed language charts + validation in Scenario_Picker
    try:
        wb = load_workbook(CFG.OUTPUT_LANG_XLSX)
        ws = wb['Meta'] if 'Meta' in wb.sheetnames else wb.active
        img4 = XLImage(fig4_path); img4.anchor = 'E2'; ws.add_image(img4)
        img5 = XLImage(fig5_path); img5.anchor = 'E22'; ws.add_image(img5)
        wb.save(CFG.OUTPUT_LANG_XLSX)
    except Exception as e:
        print(f"Chart embedding warning (language): {e}")

    # Data validation for Scenario_Picker (18/20/24)
    try:
        wb = load_workbook(CFG.OUTPUT_LANG_XLSX)
        ws = wb['Scenario_Picker']
        dv = DataValidation(
            type="list", formula1='"18,20,24"', allow_blank=True,
            showErrorMessage=True, errorTitle="Invalid", error="Use 18, 20 or 24 on weekdays"
        )
        max_row = ws.max_row
        dv.add(f"C2:C{max_row}")
        ws.add_data_validation(dv)
        ws.cell(row=1, column=4, value=(
            "Scenario_WD applies to Mon–Fri (18/20/24). Weekends fixed at 7 agents. "
            f"Language base for staffing: {lang_base_col}. Agents = ceil(workload_lang/{CFG.CASES_PER_AGENT})"
        ))
        wb.save(CFG.OUTPUT_LANG_XLSX)
    except Exception as e:
        print(f"Data validation warning: {e}")

    print(f"Main workbook: {CFG.OUTPUT_XLSX}")
    print(f"Language staffing workbook: {CFG.OUTPUT_LANG_XLSX}")

if __name__ == '__main__':
    main()